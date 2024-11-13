from flask import Flask, render_template, request, redirect, url_for, session, Response, send_file
from model_2 import prepro_img, letter_pred
from utils import stat, change_num, df_show, graphics
from werkzeug.utils import secure_filename
from PIL import Image
from io import StringIO
from io import BytesIO
import pandas as pd
import os 
import zipfile
import shutil


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


app = Flask(__name__, static_url_path='/static')
app.secret_key = "12345"

@app.route('/')
def home():
    return render_template('home.html')




@app.route('/dev_df', methods=['GET', 'POST'])
def dev_df():
    if request.method == 'POST':
        num_q = int(request.form.get('numQuestions'))
        num_o = int(request.form.get('numOptions'))
        
        # Replace with your actual logic
        options = change_num(number_options=num_o, number_questions=num_q)
        df = df_show(num_options=num_q, num_questions=num_o)

        # Store the DataFrame and other data in the session
        session['df'] = df.to_html()
        session['num_q'] = num_q
        session['num_o'] = num_o
        session['options'] = options
        
        return redirect(url_for('dev_df'))
    else:
        df_html = session.get('df', "")
        num_q = session.get('num_q', 0)
        num_o = session.get('num_o', 0)
        options = session.get('options', [])
    
    return render_template('dev_df.html', df_html=df_html, num_q=num_q, num_o=num_o, options=options)



@app.route('/download_excel')
def download_excel():
    num_q = session.get('num_q', 0)
    num_o = session.get('num_o', 0)
    df = df_show(num_options=num_o, num_questions=num_q)
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    num_options_list = ['' for _ in range(num_q)]
    num_questions_list = [1 + i for i in range(num_q)]
    df_data = {
        'P  R  E  G  U  N  T  A': num_questions_list,
        'R  E  S  P  U  E  S  T  A ': num_options_list
    }
    df = pd.DataFrame(df_data)
    df = df.transpose()

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name='Tahoma', size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')  
            cell.border = border_style  

    first_col_letter = ws.cell(row=1, column=1).column_letter
    ws.column_dimensions[first_col_letter].width = 40  

    default_width = 10  
    for c_idx, col in enumerate(df.columns, start=2):  
        col_letter = chr(64 + c_idx)  
        ws.column_dimensions[col_letter].width = default_width

    for row in ws.rows:
        row_number = row[0].row
        ws.row_dimensions[row_number].height = 30 

    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment;filename=output.xlsx"}
    )






@app.route('/submit_answers', methods=['POST'])
def submit_answers():
    num_q = int(request.form.get('num_q'))

    options = session.get('options', [])

    df = pd.DataFrame({
        'Pregunta': [ i+1 for i in range(num_q)],
        'Respuesta': [request.form.get(f'{i}') for i in range(num_q)]
    })
    print(df)
    session['df'] = df.to_html()

    answers_student = [request.form.get(f'{i}') for i in range(num_q)]
    session['answer_student']=answers_student
    print(answers_student, 20*'*')


    return redirect(url_for('dev_df'))





@app.route('/take', methods=['GET', 'POST'])
def take():
    df_html = session.get('df', None)
    return render_template('take.html', df_html=df_html)


@app.route('/prediction', methods=['GET', 'POST'])
def prediction():
    if request.method == 'POST':
        if 'imageFile' not in request.files:
            return redirect(request.url)

        file = request.files['imageFile']

        if file.filename == '':
            return redirect(request.url)

        upload_folder = './static/uploads'       
        file_path = os.path.join(upload_folder, file.filename)
    
        file_path = file_path.replace('\\', '/')
        file.save(file_path)


        upload_folder_2 = '/uploads'       
        file_path_2 = os.path.join(upload_folder_2, file.filename)
        file_path_2 = file_path_2.replace('\\', '/')    

        gray, img = prepro_img(file_path)

        answer_list = letter_pred(img, gray)
        
        answer_student = session.get('answer_student', [])


        grade_point = 100/len(answer_student)
        exam_grade = 0
        for i in range(len(answer_student)):
            if answer_student[i] == answer_list[i]:
                exam_grade += grade_point


        return render_template('prediction_result.html', image_path=file_path_2, answer=answer_list, exam_grade=int(exam_grade))

    return render_template('prediction.html')


















app.config['UPLOAD_FOLDER'] = './static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'zip'}
app.config['ALLOWED_IMAGE_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def handle_remove_readonly(func, path, exc_info):
    os.chmod(path, 0o777)  # Cambia los permisos del archivo o carpeta
    func(path) 


@app.route('/group_predictions', methods=['GET', 'POST'])
def group_predictions():
    if request.method == 'POST':
        if 'imageFile' not in request.files:
            return redirect(request.url)

        file = request.files['imageFile']

        if file.filename == '':
            return redirect(request.url)

        if not allowed_file(file.filename, app.config['ALLOWED_EXTENSIONS']):
            return redirect(request.url)
        
        filename = secure_filename(file.filename)
        print(filename, 30 * '*')
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # Unzip the file
        extracted_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'extracted')

        os.makedirs(extracted_folder, exist_ok=True)

        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(extracted_folder)

        results = []
        for root, _, files in os.walk(extracted_folder):
            print(files, 'Es una prueba')
            for file in files:
                if not allowed_file(file, app.config['ALLOWED_IMAGE_EXTENSIONS']):
                    continue

                image_path = os.path.join(root, file)

                gray, img = prepro_img(image_path)
                answer_str = letter_pred(img, gray)
                answer_list = [i for i in answer_str]

                answer_student = session.get('answer_student', [])

                grade_point = 100 / len(answer_student) if len(answer_student) > 0 else 0
                exam_grade = 0
                for i in range(len(answer_student)):
                    if answer_student[i] == answer_list[i]:
                        exam_grade += grade_point

                results.append({
                    'image_path': image_path,
                    'answer': answer_str,
                    'exam_grade': int(exam_grade)
                })

        session['data_frame'] = results
        results_df = pd.DataFrame(results)

        results_file = os.path.join(app.config['UPLOAD_FOLDER'], 'results.xlsx')
        results_df.to_excel(results_file, index=False)

        try:
            os.remove(file_path)  # Elimina el archivo .zip
        except Exception as e:
            print(f"Error al eliminar el archivo zip: {e}")

        try:
            shutil.rmtree(extracted_folder, onerror=handle_remove_readonly)  # Elimina la carpeta descomprimida
        except Exception as e:
            print(f"Error al eliminar la carpeta descomprimida: {e}")

        session['results_file'] = results_file


        return redirect(url_for('download_results'))
        # return send_file(results_file, as_attachment=True)

    return render_template('group_predictions.html')  



@app.route('/download_results')
def download_results():
    results_file = session.get('results_file')
    if results_file:
        return send_file(results_file, as_attachment=True)
    return redirect(url_for('group_predictions'))


@app.route('/statistics', methods=['GET', 'POST'])
def statistics():
    

    image_bar, image_chart, image_frec, image_chart_frec = graphics()
    mean_score, median_score, mode_score, std_score, var_score, max_score, min_score, approved_students, failed_students = stat()

    return render_template('statistics.html', 
                            mean=mean_score, 
                            median=median_score, 
                            mode = mode_score,
                            std_dev=std_score, 
                            variance=var_score, 
                            max_score = max_score,
                            min_score = min_score,
                            approved_students = approved_students, 
                            failed_students= failed_students,
                        
                            image_bar =image_bar, 
                            image_chart =image_chart,
                            image_frec = image_frec,
                            image_chart_frec =image_chart_frec)



if __name__ == '__main__':
    app.run(debug = True)


