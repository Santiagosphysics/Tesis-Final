import pandas as pd
import matplotlib.pyplot as plt 
import seaborn as sns 

# def statics(path_file):
#     df = pd.read_excel(io=path_f)
#     score_student = df['exam_grade']
#     mean_score = score_student.mean()
#     median_score = score_student.median()
#     std_score = score_student.std()
#     var_score = score_student.var()

#     plt.figure(figsize=(10,6))
#     plt.hist(score_student, bins=10, color='skyblue', edgecolor='black')
#     plt.xlabel('Score')
#     image_hist= 'hist_plot.png'
#     plt.savefig(image_hist)
#     plt.close()

#     plt.figure(figsize=(10,6))
#     plt.boxplot(score_student)
#     plt.xlabel('Score')
#     image_box= 'box_plot.png'
#     plt.savefig(image_box)
#     plt.close()


#     return mean_score, median_score, std_score, var_score, image_hist, image_box


def organize(num_option):
    num = '000'
    x = num_option
    list_1 = []
    list_2 = []
    list_3 = []

    last_num = int(num[-1])
    second_num = int(num[-2])
    first_num = int(num[0])

    for i in range(x):
        i = 1
        last_num = last_num + i
        if last_num == 10:
            second_num += 1
            last_num = 0

        if second_num == 10:
            first_num += 1
            second_num = 0
        
        list_1.append(last_num)
        list_2.append(second_num)
        list_3.append(first_num)

    list_4 = [list_3, list_2, list_1]

    df_test = pd.DataFrame(list_4).T
    response = []
    final_row = ''
    for i in range(x):
        for j in df_test.loc[i][::]:
            final_row += str(j)
        response.append(final_row)
        final_row=''

        
    return response





import pandas as pd
import matplotlib.pyplot as plt
import os

def stat():

    path_f = r'static/uploads/results.xlsx'
    path_f = path_f.replace('\\', '/')
    
    df = pd.read_excel(io=path_f)
    score_student = df['exam_grade']

    mean_score = score_student.mean()
    median_score = score_student.median()
    std_score = score_student.std()
    var_score = score_student.var()
    max_score = score_student.max()
    min_score = score_student.min()

    dict_1 = {}
    for i in score_student:
        if i not in dict_1:
            dict_1[i] = 1
        else:
            dict_1[i] += 1
    mode_score = next( key for key, value in dict_1.items() if value == max(dict_1.values()))

    approved_students = 0
    failed_students = 0
    
    for i in score_student:
        if i >= 70:
            approved_students+=1 
        else:
            failed_students += 1

    return mean_score, median_score, mode_score, std_score, var_score, max_score, min_score, approved_students, failed_students

def graphics():
    path_f = r'static/uploads/results.xlsx'
    path_f = path_f.replace('\\', '/')
    
    df = pd.read_excel(io=path_f)
    score_student = df['exam_grade']
    os.makedirs('static/uploads', exist_ok=True)  
    colors = ['green' if score >= 70 else 'red' for score in score_student ]
    plt.figure(figsize=(10,6))
    plt.bar(x= [i+1 for i in range(len(score_student))], height= score_student, color=colors)
    plt.title('Grade and students')
    plt.xlabel('List number')
    plt.ylabel('Score')
    image_bar = 'static/uploads/bar_plot.png'
    plt.savefig(image_bar)
    plt.close()



    dict_2 = {}
    for i in score_student:
        if i  in dict_2:
            dict_2[i] += 1
        else:
            dict_2[i] = 1
        
    scores = list(dict_2.keys())
    frec_answ = list(dict_2.values())

    colors_2 = ['green' if score >= 70 else 'red' for score in scores]
    plt.figure(figsize=(10,6))
    plt.bar(x = scores, height=frec_answ, color=colors_2)
    plt.title('Frecuency and score')
    plt.xlabel('Score')
    plt.ylabel('Frecuency')
    image_frec = 'static/uploads/bar_fre_plot.png'
    plt.savefig(image_frec)
    plt.close()

    plt.figure(figsize=(10,6))
    plt.pie(x=frec_answ, labels= [f'{i} points' for i in scores] , startangle=0, autopct='%1.1f%%')
    plt.title('Frecuency and score')
    image_chart_frec = 'static/uploads/pie_chart_frec.png'
    plt.savefig(image_chart_frec)
    plt.close()
    
    plt.figure(figsize=(10,6))
    plt.pie(x=score_student, labels = [f'Student {i+1} ' for i in range(len(score_student))])
    image_chart = 'static/uploads/pie_chart.png'
    plt.savefig(image_chart)
    plt.close()


    return image_bar, image_chart, image_frec, image_chart_frec




import numpy as np 
import pandas as pd 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def change_num(number_options, number_questions):
    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    options =  [abc[i] for i in range(number_options)]

    return options



def creation_test(number_questions, number_options, options):
    df = { i+1:[ 'O' for _ in range(number_options)] for i in range(number_questions)}
    df = pd.DataFrame(df)
    options = change_num(number_options, number_questions)
    df.index = options
    return df


# def df_show(num_options, num_questions):
#     num_options_list = ['' for i in range(num_options)]
#     num_questions_list = [1+i for i in range(num_options)]

#     df = {
#         'P R E G U N T A': num_questions_list,
#         'R E S P U E S T A ': num_options_list
#         }
    
#     df = pd.DataFrame(df)
#     df = df.transpose()
#     # df = df.to_csv('df.csv', header=False)
#     return df



def df_show(num_options, num_questions):
    num_options_list = ['' for _ in range(num_questions)]
    num_questions_list = [1 + i for i in range(num_questions)]

    df = {
        'P  R  E  G  U  N  T  A': num_questions_list,
        'R  E  S  P  U  E  S  T  A ': num_options_list
    }
    df = pd.DataFrame(df)
    df = df.transpose()

    wb = Workbook()
    ws = wb.active

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name='Tahoma', size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')  
            cell.border = border_style  

    first_col_letter = ws.cell(row=1, column=1).column_letter
    ws.column_dimensions[first_col_letter].width = 40  

    default_width = 10  
    for c_idx, col in enumerate(df.columns, start=2):  
        col_letter = chr(64 + c_idx)  
        ws.column_dimensions[col_letter].width = default_width

    for row in ws.rows:
        row_number = row[0].row
        ws.row_dimensions[row_number].height = 30 

    return df